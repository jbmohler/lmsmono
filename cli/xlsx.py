"""Shared xlsx building blocks for consistently-themed reports."""

import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Shared constants — import these by name in report modules
# ---------------------------------------------------------------------------

REPORT_FONT = "Nimbus Sans"

FILL_SECTION = PatternFill("solid", fgColor="D9E1F2")   # blue-gray section header
FILL_SUBTOTAL = PatternFill("solid", fgColor="E2EFDA")  # green subtotal
FILL_ASSETS = PatternFill("solid", fgColor="BDD7EE")    # blue  — total assets
FILL_LIAB = PatternFill("solid", fgColor="FCE4D6")      # peach — total liabilities
FILL_GROUP = PatternFill("solid", fgColor="EBEBEB")     # light gray sub-group header

CURRENCY_FMT = '#,##0.00;[Red]-#,##0.00'

# ---------------------------------------------------------------------------
# ReportSheet
# ---------------------------------------------------------------------------


class ReportSheet:
    """Stateful row-cursor builder for a single consistently-styled worksheet.

    Usage pattern:
        rs = ReportSheet("My Report", n_text=2, n_val=3)
        rs.write_title("My Report")
        rs.write_generated()
        rs.write_blank()
        rs.write_headers(["Account", "Journal"], ["Jan", "Feb", "Mar"])
        rs.set_col_widths([28, 20, 16, 16, 16])
        # ... write section headers, data rows, subtotal rows ...
        rs.save("report.xlsx")
    """

    def __init__(self, sheet_title: str, n_text: int, n_val: int) -> None:
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = sheet_title
        self.n_text = n_text
        self.n_val = n_val
        self.total_cols = n_text + n_val
        self.val_col = n_text + 1   # 1-based index of first value column
        self.row = 1

        self._title_font = Font(name=REPORT_FONT, bold=True, size=14)
        self._gray_font = Font(name=REPORT_FONT, color="808080", size=10)
        self._header_font = Font(name=REPORT_FONT, bold=True, size=11)
        self._section_font = Font(name=REPORT_FONT, bold=True, size=11)

    # ------------------------------------------------------------------
    # Preamble
    # ------------------------------------------------------------------

    def write_title(self, text: str) -> None:
        """Row: bold large title spanning all columns."""
        self._merge_row(self.row)
        c = self.ws.cell(self.row, 1, text)
        c.font = self._title_font
        c.alignment = Alignment(vertical="center")
        self.ws.row_dimensions[self.row].height = 28
        self.row += 1

    def write_generated(self) -> None:
        """Row: gray 'Generated YYYY-MM-DD HH:MM' spanning all columns."""
        self._merge_row(self.row)
        stamp = datetime.datetime.now().strftime("Generated %Y-%m-%d %H:%M")
        self.ws.cell(self.row, 1, stamp).font = self._gray_font
        self.row += 1

    def write_note(self, text: str) -> None:
        """Row: gray italic note spanning all columns."""
        self._merge_row(self.row)
        self.ws.cell(self.row, 1, text).font = self._gray_font
        self.row += 1

    def write_blank(self) -> None:
        """Advance one blank row."""
        self.row += 1

    def write_period_header(self, label: str) -> None:
        """Row: italic label merged across value columns only (text columns left blank)."""
        if self.n_val > 1:
            self.ws.merge_cells(
                start_row=self.row, start_column=self.val_col,
                end_row=self.row, end_column=self.val_col + self.n_val - 1,
            )
        self.ws.cell(self.row, self.val_col, label).font = Font(name=REPORT_FONT, italic=True, color="808080", size=10)
        self.row += 1

    def write_headers(self, text_headers: list[str], val_headers: list[str]) -> None:
        """Row: bold column headers; freeze pane set below this row."""
        vcenter = Alignment(vertical="center")
        hcenter = Alignment(horizontal="center", vertical="center")
        for col, label in enumerate(text_headers, start=1):
            c = self.ws.cell(self.row, col, label)
            c.font = self._header_font
            c.alignment = vcenter
        for i, label in enumerate(val_headers):
            c = self.ws.cell(self.row, self.val_col + i, label)
            c.font = self._header_font
            c.alignment = hcenter
        self.ws.row_dimensions[self.row].height = 20
        self.ws.freeze_panes = f"A{self.row + 1}"
        self.row += 1

    # ------------------------------------------------------------------
    # Section and data rows
    # ------------------------------------------------------------------

    def write_section_header(self, label: str) -> None:
        """Row: full-width merged section label with section fill."""
        self._merge_row(self.row)
        c = self.ws.cell(self.row, 1, label)
        c.font = self._section_font
        c.fill = FILL_SECTION
        c.alignment = Alignment(vertical="center")
        self.ws.row_dimensions[self.row].height = 20
        self.row += 1

    def write_group_header(self, texts: list[str]) -> None:
        """Row: lightly shaded sub-group label within a section (e.g. account name)."""
        self._fill_row(FILL_GROUP)
        for col, text in enumerate(texts, start=1):
            self.ws.cell(self.row, col, text).font = Font(name=REPORT_FONT, bold=True, size=10)
        self.row += 1

    def write_data_row(self, texts: list[str], values: list[float]) -> int:
        """Row: text cells then right-aligned currency cells. Returns row number."""
        for col, val in enumerate(texts, start=1):
            self.ws.cell(self.row, col, val)
        for i, val in enumerate(values):
            c = self.ws.cell(self.row, self.val_col + i, val)
            c.number_format = CURRENCY_FMT
            c.alignment = Alignment(horizontal="right")
        written = self.row
        self.row += 1
        return written

    # ------------------------------------------------------------------
    # Formula rows
    # ------------------------------------------------------------------

    def write_subtotal_row(
        self,
        label: str,
        first_data_row: int,
        fill: PatternFill = FILL_SUBTOTAL,
    ) -> int:
        """Row: =SUM over [first_data_row .. row-1]. Returns row number."""
        self._fill_row(fill)
        self.ws.cell(self.row, 1, label).font = Font(name=REPORT_FONT, bold=True)
        for i in range(self.n_val):
            col = get_column_letter(self.val_col + i)
            c = self.ws.cell(self.row, self.val_col + i, f"=SUM({col}{first_data_row}:{col}{self.row - 1})")
            c.number_format = CURRENCY_FMT
            c.font = Font(name=REPORT_FONT, bold=True)
            c.alignment = Alignment(horizontal="right")
        written = self.row
        self.row += 1
        return written

    def write_ref_row(
        self,
        label: str,
        ref_rows: list[int],
        fill: PatternFill = FILL_SUBTOTAL,
    ) -> int:
        """Row: values are the sum of specific ref_rows. Returns row number."""
        return self.write_signed_ref_row(label, add_rows=ref_rows, sub_rows=[], fill=fill)

    def write_signed_ref_row(
        self,
        label: str,
        add_rows: list[int],
        sub_rows: list[int],
        fill: PatternFill = FILL_SUBTOTAL,
    ) -> int:
        """Row: values = sum(add_rows) - sum(sub_rows). Returns row number."""
        self._fill_row(fill)
        self.ws.cell(self.row, 1, label).font = Font(name=REPORT_FONT, bold=True, size=11)
        for i in range(self.n_val):
            col = get_column_letter(self.val_col + i)
            parts = [f"{col}{r}" for r in add_rows] + [f"-{col}{r}" for r in sub_rows]
            formula = ("=" + "+".join(parts)) if parts else "=0"
            c = self.ws.cell(self.row, self.val_col + i, formula)
            c.number_format = CURRENCY_FMT
            c.font = Font(name=REPORT_FONT, bold=True, size=11)
            c.alignment = Alignment(horizontal="right")
        written = self.row
        self.row += 1
        return written

    # ------------------------------------------------------------------
    # Finishing
    # ------------------------------------------------------------------

    def set_col_widths(self, widths: list[float]) -> None:
        """Set column widths; list is 1-based (index 0 → column A)."""
        for i, w in enumerate(widths, start=1):
            self.ws.column_dimensions[get_column_letter(i)].width = w

    def save(self, path: str) -> None:
        self.wb.save(path)

    # ------------------------------------------------------------------
    # Internals
    # ------------------------------------------------------------------

    def _merge_row(self, row: int) -> None:
        self.ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row, end_column=self.total_cols,
        )

    def _fill_row(self, fill: PatternFill) -> None:
        for col in range(1, self.total_cols + 1):
            self.ws.cell(self.row, col).fill = fill
