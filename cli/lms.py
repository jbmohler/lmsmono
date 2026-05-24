#!/usr/bin/env python3
"""LMS command-line client."""

import argparse
import datetime
import getpass
import json
import os
import sys

import httpx
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

CONFIG_DIR = os.path.join(os.path.expanduser("~"), ".lms")
SESSION_FILE = os.path.join(CONFIG_DIR, "session")
DEFAULT_BASE_URL = "http://localhost:8080"


# ---------------------------------------------------------------------------
# Session management
# ---------------------------------------------------------------------------


def load_session() -> dict:
    if os.path.exists(SESSION_FILE):
        try:
            with open(SESSION_FILE) as f:
                return json.loads(f.read())
        except Exception:
            return {}
    return {}


def save_session(data: dict) -> None:
    os.makedirs(CONFIG_DIR, exist_ok=True)
    with open(SESSION_FILE, "w") as f:
        f.write(json.dumps(data))
    os.chmod(SESSION_FILE, 0o600)


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------


def _do_login(base_url: str) -> str:
    """Prompt for credentials, POST to login, return session_id."""
    print("Authentication required.")
    username = input("Username: ").strip()
    password = getpass.getpass("Password: ")

    with httpx.Client(base_url=base_url) as client:
        resp = client.post(
            "/api/auth/login",
            json={"username": username, "password": password},
        )

    if resp.status_code not in (200, 201):
        try:
            detail = resp.json().get("detail", "Login failed")
        except Exception:
            detail = f"HTTP {resp.status_code}"
        print(f"Error: {detail}", file=sys.stderr)
        sys.exit(1)

    session_id = resp.cookies.get("session_id")
    if not session_id:
        print("Error: No session cookie in response", file=sys.stderr)
        sys.exit(1)

    return session_id


def ensure_auth(base_url: str | None) -> httpx.Client:
    """Return an authenticated HTTP client, logging in if necessary.

    Checks the saved session first; re-authenticates on 401.
    """
    session = load_session()
    effective_url = base_url or session.get("base_url") or DEFAULT_BASE_URL

    if session.get("session_id"):
        client = httpx.Client(
            base_url=effective_url,
            cookies={"session_id": session["session_id"]},
        )
        resp = client.get("/api/auth/me")
        if resp.status_code == 200:
            return client
        client.close()

    session_id = _do_login(effective_url)
    save_session({"session_id": session_id, "base_url": effective_url})
    print("Logged in.")
    return httpx.Client(
        base_url=effective_url,
        cookies={"session_id": session_id},
    )


# ---------------------------------------------------------------------------
# Balance sheet export
# ---------------------------------------------------------------------------


def cmd_balance_sheet(args: argparse.Namespace) -> None:
    today = datetime.date.today()
    year: int = args.year or today.year
    month: int = args.month or today.month
    periods: int = args.periods

    client = ensure_auth(args.url)
    resp = client.get(
        "/api/reports/multi-period-balance-sheet",
        params={"year": year, "month": month, "periods": periods},
    )
    client.close()

    if resp.status_code != 200:
        try:
            detail = resp.json().get("detail", f"HTTP {resp.status_code}")
        except Exception:
            detail = f"HTTP {resp.status_code}"
        print(f"Error: {detail}", file=sys.stderr)
        sys.exit(1)

    payload = resp.json()
    period_dates: list[str] = payload["periods"]
    rows: list[dict] = payload["data"]

    output: str = args.output or f"balance_sheet_{year}_{month:02d}.xlsx"
    _write_balance_sheet_xlsx(output, period_dates, rows)
    print(f"Written: {output}")


def _write_balance_sheet_xlsx(
    path: str,
    period_dates: list[str],
    rows: list[dict],
) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"

    n = len(period_dates)
    total_cols = 2 + n

    # Styles
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11)
    section_font = Font(bold=True, size=11)
    section_fill = PatternFill("solid", fgColor="D9E1F2")
    subtotal_fill = PatternFill("solid", fgColor="E2EFDA")
    grand_fill = PatternFill("solid", fgColor="FFF2CC")
    currency_fmt = '#,##0.00;[Red]-#,##0.00'
    right = Alignment(horizontal="right")
    center = Alignment(horizontal="center")

    # Row 1: title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws.cell(1, 1, "Balance Sheet")
    c.font = title_font

    # Row 2: column headers
    ws.cell(2, 1, "Journal").font = header_font
    ws.cell(2, 2, "Account").font = header_font
    for i, d in enumerate(period_dates):
        c = ws.cell(2, 3 + i, d)
        c.font = header_font
        c.alignment = center

    ws.freeze_panes = "A3"

    row_num = 3
    current_atype: str | None = None
    section_totals: list[float] = [0.0] * n
    grand_totals: list[float] = [0.0] * n

    def flush_section() -> None:
        nonlocal row_num
        if current_atype is None:
            return
        ws.cell(row_num, 1, f"Total {current_atype}").font = Font(bold=True)
        ws.cell(row_num, 1).fill = subtotal_fill
        ws.cell(row_num, 2).fill = subtotal_fill
        for i, val in enumerate(section_totals):
            c = ws.cell(row_num, 3 + i, val)
            c.number_format = currency_fmt
            c.font = Font(bold=True)
            c.fill = subtotal_fill
            c.alignment = right
        row_num += 1

    for data_row in rows:
        atype = data_row["atype_name"]

        if atype != current_atype:
            flush_section()
            current_atype = atype
            section_totals = [0.0] * n
            # Section header spanning all columns
            ws.merge_cells(
                start_row=row_num, start_column=1,
                end_row=row_num, end_column=total_cols,
            )
            c = ws.cell(row_num, 1, atype)
            c.font = section_font
            c.fill = section_fill
            row_num += 1

        ws.cell(row_num, 1, data_row["journal"]["name"])
        ws.cell(row_num, 2, data_row["acc_name"])
        for i, bal in enumerate(data_row["balances"]):
            val = float(bal or 0.0)
            c = ws.cell(row_num, 3 + i, val)
            c.number_format = currency_fmt
            c.alignment = right
            section_totals[i] += val
            grand_totals[i] += val

        row_num += 1

    flush_section()

    # Grand total row
    ws.cell(row_num, 1, "TOTAL").font = Font(bold=True, size=11)
    ws.cell(row_num, 1).fill = grand_fill
    ws.cell(row_num, 2).fill = grand_fill
    for i, val in enumerate(grand_totals):
        c = ws.cell(row_num, 3 + i, val)
        c.number_format = currency_fmt
        c.font = Font(bold=True, size=11)
        c.alignment = right
        c.fill = grand_fill

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 32
    for i in range(n):
        ws.column_dimensions[get_column_letter(3 + i)].width = 16

    wb.save(path)


# ---------------------------------------------------------------------------
# Logout
# ---------------------------------------------------------------------------


def cmd_logout(args: argparse.Namespace) -> None:
    session = load_session()
    effective_url = args.url or session.get("base_url") or DEFAULT_BASE_URL

    if session.get("session_id"):
        with httpx.Client(
            base_url=effective_url,
            cookies={"session_id": session["session_id"]},
        ) as client:
            client.post("/api/auth/logout")

    if os.path.exists(SESSION_FILE):
        os.remove(SESSION_FILE)
    print("Logged out.")


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(
        prog="lms",
        description="LMS command-line client",
    )
    parser.add_argument(
        "--url",
        default=None,
        metavar="URL",
        help=f"Base URL of the LMS server (default: {DEFAULT_BASE_URL})",
    )

    subparsers = parser.add_subparsers(dest="command", required=True)

    # balance-sheet command
    bs = subparsers.add_parser(
        "balance-sheet",
        help="Export balance sheet to Excel",
    )
    bs.add_argument("--year", type=int, default=None, help="Report year (default: current year)")
    bs.add_argument("--month", type=int, default=None, help="Report month 1-12 (default: current month)")
    bs.add_argument("--periods", type=int, default=3, metavar="N", help="Number of annual periods (default: 3)")
    bs.add_argument("--output", "-o", default=None, metavar="FILE", help="Output .xlsx file (default: balance_sheet_YYYY_MM.xlsx)")
    bs.set_defaults(func=cmd_balance_sheet)

    # logout command
    lo = subparsers.add_parser("logout", help="Clear saved session")
    lo.set_defaults(func=cmd_logout)

    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
